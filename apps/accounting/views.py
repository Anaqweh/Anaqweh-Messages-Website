from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer


@login_required
def _acc_tenant(request):
    """شركة المستخدم الحالي للعزل (المدير العام = None يرى الكل)."""
    if request.user.is_superuser:
        return None
    try:
        from apps.payments.views import _finance_tenant_for_request
        return _finance_tenant_for_request(request)
    except Exception:
        return None

def _acc_customers(request):
    qs = Customer.objects.all()
    _t = _acc_tenant(request)
    if _t is not None:
        qs = qs.filter(tenant=_t)
    elif not request.user.is_superuser:
        qs = qs.none()
    return qs

def customers(request):
    q = (request.GET.get('q') or '').strip()
    qs = _acc_customers(request)
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q) | Q(company__icontains=q))
    return render(request, 'accounting/customers.html', {'customers': qs, 'q': q})


@login_required
def customer_create(request):
    if request.method == 'POST':
        Customer.objects.create(
            tenant=_acc_tenant(request),
            name=request.POST.get('name', '').strip() or 'عميل',
            email=request.POST.get('email', '').strip(),
            phone=request.POST.get('phone', '').strip(),
            company=request.POST.get('company', '').strip(),
            address=request.POST.get('address', '').strip(),
            trn=request.POST.get('trn', '').strip(),
            notes=request.POST.get('notes', '').strip(),
        )
        return redirect('accounting:customers')
    return render(request, 'accounting/customer_form.html', {'customer': None})


@login_required
def customer_edit(request, pk):
    c = get_object_or_404(_acc_customers(request), pk=pk)
    if request.method == 'POST':
        c.name = request.POST.get('name', '').strip() or c.name
        c.email = request.POST.get('email', '').strip()
        c.phone = request.POST.get('phone', '').strip()
        c.company = request.POST.get('company', '').strip()
        c.address = request.POST.get('address', '').strip()
        c.trn = request.POST.get('trn', '').strip()
        c.notes = request.POST.get('notes', '').strip()
        c.is_active = bool(request.POST.get('is_active'))
        c.save()
        return redirect('accounting:customer_detail', pk=c.pk)
    return render(request, 'accounting/customer_form.html', {'customer': c})


@login_required
def customer_detail(request, pk):
    c = get_object_or_404(_acc_customers(request), pk=pk)
    invoices = c._sales_invoices().prefetch_related('items')
    return render(request, 'accounting/customer_detail.html', {'customer': c, 'invoices': invoices})


def _report_data(request):
    """يجمع بيانات التقرير - قراءة فقط من جداول الدفع/الفواتير."""
    from decimal import Decimal
    from django.utils import timezone
    from apps.payments.models import SalesInvoice, Expense
    today = timezone.localdate()
    start = request.GET.get('from', '')
    end = request.GET.get('to', '')
    invoices = SalesInvoice.objects.all()
    expenses = Expense.objects.all()
    # عزل حسب الشركة (المدير العام يرى الكل)
    try:
        from apps.payments.views import _finance_tenant_for_request
        _ft = _finance_tenant_for_request(request)
        if _ft is not None:
            invoices = invoices.filter(tenant=_ft)
            expenses = expenses.filter(tenant=_ft)
        elif not request.user.is_superuser:
            invoices = invoices.none()
            expenses = expenses.none()
    except Exception:
        pass
    if start:
        invoices = invoices.filter(issue_date__gte=start)
        expenses = expenses.filter(spent_at__gte=start)
    if end:
        invoices = invoices.filter(issue_date__lte=end)
        expenses = expenses.filter(spent_at__lte=end)
    invoices = list(invoices.prefetch_related('items'))
    paid = [i for i in invoices if i.status == 'paid']
    unpaid = [i for i in invoices if i.status == 'unpaid']
    revenue = sum((i.subtotal for i in paid), Decimal('0.00'))
    tax_collected = sum((i.tax_amount for i in paid), Decimal('0.00'))
    gross = sum((i.total for i in paid), Decimal('0.00'))
    total_expenses = sum((e.amount for e in expenses), Decimal('0.00'))
    net_profit = gross - total_expenses
    outstanding = sum((i.total for i in unpaid), Decimal('0.00'))
    return {
        'today': today, 'start': start, 'end': end,
        'invoices': invoices, 'paid': paid, 'unpaid': unpaid,
        'expenses': list(expenses),
        'revenue': revenue, 'tax_collected': tax_collected, 'gross': gross,
        'total_expenses': total_expenses, 'net_profit': net_profit,
        'outstanding': outstanding,
        'cnt_paid': len(paid), 'cnt_unpaid': len(unpaid), 'cnt_total': len(invoices),
    }


@login_required
def reports(request):
    data = _report_data(request)
    return render(request, 'accounting/reports.html', data)


def _export_rows(data):
    """صفوف موحّدة للتصدير."""
    header = ['رقم الفاتورة', 'النوع', 'العميل', 'الحالة', 'الفرعي', 'الضريبة', 'الإجمالي', 'العملة', 'تاريخ الإصدار']
    rows = []
    for i in data['invoices']:
        rows.append([
            i.invoice_number, i.get_kind_display(), i.customer_name,
            i.get_status_display(), float(i.subtotal), float(i.tax_amount),
            float(i.total), i.currency.upper(), i.issue_date.strftime('%Y-%m-%d'),
        ])
    return header, rows


@login_required
def reports_export_csv(request):
    import csv
    from django.http import HttpResponse
    data = _report_data(request)
    header, rows = _export_rows(data)
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = 'attachment; filename="financial_report.csv"'
    resp.write('\ufeff')  # BOM لدعم العربية في Excel
    writer = csv.writer(resp)
    writer.writerow(header)
    writer.writerows(rows)
    writer.writerow([])
    writer.writerow(['الإيرادات', float(data['revenue'])])
    writer.writerow(['الضريبة المحصّلة', float(data['tax_collected'])])
    writer.writerow(['إجمالي المبيعات', float(data['gross'])])
    writer.writerow(['المصروفات', float(data['total_expenses'])])
    writer.writerow(['صافي الربح', float(data['net_profit'])])
    writer.writerow(['المستحقات القائمة', float(data['outstanding'])])
    return resp


@login_required
def reports_export_xlsx(request):
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    data = _report_data(request)
    header, rows = _export_rows(data)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'التقرير المالي'
    ws.sheet_view.rightToLeft = True
    hf = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='0B4EA2')
    ws.append(header)
    for c in ws[1]:
        c.font = hf
        c.fill = fill
    for r in rows:
        ws.append(r)
    ws.append([])
    summary = [
        ('الإيرادات', float(data['revenue'])),
        ('الضريبة المحصّلة', float(data['tax_collected'])),
        ('إجمالي المبيعات', float(data['gross'])),
        ('المصروفات', float(data['total_expenses'])),
        ('صافي الربح', float(data['net_profit'])),
        ('المستحقات القائمة', float(data['outstanding'])),
    ]
    for label, val in summary:
        ws.append([label, val])
        ws[ws.max_row][0].font = Font(bold=True)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
    return resp


# ============================================================
# نظام رواتب الموظفين (HR + Payroll)
# ============================================================
from .models import Employee, PayrollRun, Payslip
from decimal import Decimal


def _acc_employees(request):
    """عزل الموظفين بالشركة - نفس آلية العملاء."""
    qs = Employee.objects.all()
    _t = _acc_tenant(request)
    if _t is not None:
        qs = qs.filter(tenant=_t)
    elif not request.user.is_superuser:
        qs = qs.none()
    return qs


def _acc_payrolls(request):
    qs = PayrollRun.objects.all()
    _t = _acc_tenant(request)
    if _t is not None:
        qs = qs.filter(tenant=_t)
    elif not request.user.is_superuser:
        qs = qs.none()
    return qs


@login_required
def employees(request):
    """قائمة الموظفين مع تنبيهات الهويات."""
    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or '').strip()
    qs = _acc_employees(request)
    if status:
        qs = qs.filter(status=status)
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(full_name__icontains=q) | Q(national_id__icontains=q) | Q(job_title__icontains=q) | Q(phone__icontains=q))

    # تنبيهات الهويات
    expired_count = sum(1 for e in qs if e.id_status == 'expired')
    soon_count = sum(1 for e in qs if e.id_status == 'soon')
    active_count = qs.filter(status='active').count()
    total_payroll = sum((e.net_salary for e in qs if e.status == 'active'), Decimal('0'))

    return render(request, 'accounting/employees.html', {
        'employees': qs,
        'q': q,
        'status': status,
        'expired_count': expired_count,
        'soon_count': soon_count,
        'active_count': active_count,
        'total_payroll': total_payroll,
    })


@login_required
def employee_create(request):
    if request.method == 'POST':
        Employee.objects.create(
            tenant=_acc_tenant(request),
            full_name=request.POST.get('full_name', '').strip() or 'موظف',
            national_id=request.POST.get('national_id', '').strip(),
            id_expiry=request.POST.get('id_expiry') or None,
            job_title=request.POST.get('job_title', '').strip(),
            hire_date=request.POST.get('hire_date') or None,
            base_salary=request.POST.get('base_salary') or 0,
            allowances=request.POST.get('allowances') or 0,
            deductions=request.POST.get('deductions') or 0,
            iban=request.POST.get('iban', '').strip(),
            phone=request.POST.get('phone', '').strip(),
            email=request.POST.get('email', '').strip(),
            status=request.POST.get('status', 'active'),
            notes=request.POST.get('notes', '').strip(),
        )
        from django.contrib import messages
        messages.success(request, 'تم إضافة الموظف بنجاح')
        return redirect('accounting:employees')
    return render(request, 'accounting/employee_form.html', {'employee': None})


@login_required
def employee_edit(request, pk):
    e = get_object_or_404(_acc_employees(request), pk=pk)
    if request.method == 'POST':
        e.full_name = request.POST.get('full_name', '').strip() or e.full_name
        e.national_id = request.POST.get('national_id', '').strip()
        e.id_expiry = request.POST.get('id_expiry') or None
        e.job_title = request.POST.get('job_title', '').strip()
        e.hire_date = request.POST.get('hire_date') or None
        e.base_salary = request.POST.get('base_salary') or 0
        e.allowances = request.POST.get('allowances') or 0
        e.deductions = request.POST.get('deductions') or 0
        e.iban = request.POST.get('iban', '').strip()
        e.phone = request.POST.get('phone', '').strip()
        e.email = request.POST.get('email', '').strip()
        e.status = request.POST.get('status', 'active')
        e.notes = request.POST.get('notes', '').strip()
        e.save()
        from django.contrib import messages
        messages.success(request, 'تم تحديث بيانات الموظف')
        return redirect('accounting:employees')
    return render(request, 'accounting/employee_form.html', {'employee': e})


@login_required
def employee_delete(request, pk):
    e = get_object_or_404(_acc_employees(request), pk=pk)
    if request.method == 'POST':
        e.delete()
        from django.contrib import messages
        messages.success(request, 'تم حذف الموظف')
    return redirect('accounting:employees')


@login_required
def payrolls(request):
    """قائمة مسيرات الرواتب."""
    qs = _acc_payrolls(request)
    return render(request, 'accounting/payrolls.html', {'payrolls': qs})


@login_required
def payroll_create(request):
    """توليد مسير رواتب لشهر معين تلقائياً لكل الموظفين النشطين."""
    from django.contrib import messages
    if request.method == 'POST':
        try:
            year = int(request.POST.get('year'))
            month = int(request.POST.get('month'))
        except (ValueError, TypeError):
            messages.error(request, 'الشهر والسنة مطلوبان')
            return redirect('accounting:payrolls')

        _t = _acc_tenant(request)
        # منع التكرار لنفس الشهر
        existing = _acc_payrolls(request).filter(year=year, month=month).first()
        if existing:
            messages.warning(request, 'يوجد مسير لهذا الشهر بالفعل')
            return redirect('accounting:payroll_detail', pk=existing.pk)

        run = PayrollRun.objects.create(tenant=_t, year=year, month=month, status='draft')
        # توليد قسيمة لكل موظف نشط
        active_emps = _acc_employees(request).filter(status='active')
        for emp in active_emps:
            Payslip.objects.create(
                payroll_run=run,
                employee=emp,
                base_salary=emp.base_salary or 0,
                allowances=emp.allowances or 0,
                deductions=emp.deductions or 0,
            )
        run.recalculate_total()
        messages.success(request, f'تم توليد مسير {run.month_name} {year} لـ {active_emps.count()} موظف')
        return redirect('accounting:payroll_detail', pk=run.pk)
    return redirect('accounting:payrolls')


@login_required
def payroll_detail(request, pk):
    run = get_object_or_404(_acc_payrolls(request), pk=pk)
    payslips = run.payslips.select_related('employee')
    return render(request, 'accounting/payroll_detail.html', {'run': run, 'payslips': payslips})


@login_required
def payroll_approve(request, pk):
    """اعتماد المسير + تسجيله كمصروف في المحاسبة."""
    from django.contrib import messages
    run = get_object_or_404(_acc_payrolls(request), pk=pk)
    if request.method == 'POST' and run.status == 'draft':
        run.recalculate_total()
        run.status = 'approved'
        # تسجيل مصروف تلقائي في المحاسبة
        try:
            from apps.payments.models import Expense
            import datetime
            # آخر يوم آمن في الشهر (28 يصلح لكل الشهور)
            exp = Expense.objects.create(
                tenant=run.tenant,
                title=f'رواتب الموظفين - {run.month_name} {run.year}',
                amount=run.total_amount,
                category='other',
                spent_at=datetime.date(run.year, run.month, 28),
                notes=f'مسير رواتب تلقائي ({run.payslips.count()} موظف)',
            )
            run.expense_id = exp.id
        except Exception as ex:
            messages.warning(request, f'تم الاعتماد لكن تعذّر ربط المصروف: {ex}')
        run.save()
        messages.success(request, f'تم اعتماد مسير {run.month_name} {run.year}')
    return redirect('accounting:payroll_detail', pk=pk)


@login_required
def payroll_delete(request, pk):
    from django.contrib import messages
    run = get_object_or_404(_acc_payrolls(request), pk=pk)
    if request.method == 'POST':
        run.delete()
        messages.success(request, 'تم حذف المسير')
    return redirect('accounting:payrolls')
